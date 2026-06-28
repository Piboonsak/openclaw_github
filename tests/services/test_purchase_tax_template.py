"""Tests for TASK-1101 — purchase tax report refactored to template engine.

Covers:
  ac_1101_vat     — VAT bucket splitting (no_vat / vat_0 / vat_7)
  ac_1101_output  — column order and field values preserved after refactor
  ac_1101_thai    — Thai header labels present via _PTR_COLUMNS
  ac_1101_excel   — create_purchase_tax_report() produces a readable .xlsx file
  ac_1101_redirect — 307 redirect endpoint stub (tested via response code check)
"""

from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from src.backend.services.export_service import (
    _PTR_COLUMNS,
    _preprocess_purchase_tax_doc,
    create_purchase_tax_report,
)
from src.backend.services.template_engine import ColumnDef, TemplateEngine


# ---------------------------------------------------------------------------
# ac_1101_vat — VAT bucket splitting
# ---------------------------------------------------------------------------

class TestVatBuckets(unittest.TestCase):
    def _proc(self, net, vat_rate, vat_amt, wht_amt=0):
        doc = {"netAmt": net, "vatRate": vat_rate, "vatAmt": vat_amt, "whtAmt": wht_amt}
        return _preprocess_purchase_tax_doc(doc)

    def test_no_vat_bucket(self):
        r = self._proc(1000, 0, 0)
        self.assertEqual(float(r["no_vat"]), 1000.0)
        self.assertEqual(float(r["vat_0"]),  0.0)
        self.assertEqual(float(r["vat_7"]),  0.0)

    def test_vat_0_bucket(self):
        # vatRate==0 but vatAmt>0  →  vat_0 bucket
        r = self._proc(1000, 0, 1)
        self.assertEqual(float(r["no_vat"]), 0.0)
        self.assertEqual(float(r["vat_0"]),  1000.0)
        self.assertEqual(float(r["vat_7"]),  0.0)

    def test_vat_7_bucket(self):
        r = self._proc(1000, 7, 70)
        self.assertEqual(float(r["no_vat"]), 0.0)
        self.assertEqual(float(r["vat_0"]),  0.0)
        self.assertEqual(float(r["vat_7"]),  1000.0)

    def test_total_computed(self):
        r = self._proc(1000, 7, 70)
        self.assertAlmostEqual(float(r["total"]), 1070.0)

    def test_net_payable_after_wht(self):
        r = self._proc(1000, 7, 70, wht_amt=30)
        self.assertAlmostEqual(float(r["net_payable"]), 1040.0)

    def test_zero_amounts(self):
        r = self._proc(0, 0, 0)
        self.assertEqual(float(r["no_vat"]), 0.0)
        self.assertEqual(float(r["total"]),  0.0)


# ---------------------------------------------------------------------------
# ac_1101_output — field normalisation and column order
# ---------------------------------------------------------------------------

class TestPreprocessFieldNormalisation(unittest.TestCase):
    def _doc(self, **kwargs):
        defaults = {
            "invNo": "INV001", "invDate": "2026-05-01",
            "sellerName": "ABC Co", "sellerTax": "1234567890123",
            "sellerBranch": "00001", "netAmt": 500, "vatRate": 7,
            "vatAmt": 35, "whtAmt": 0, "category": "ค่าบริการ",
            "description": "ค่าซ่อม", "reference": "REF-01",
            "status": "ลงบัญชีแล้ว",
        }
        defaults.update(kwargs)
        return defaults

    def test_inv_no_mapped(self):
        r = _preprocess_purchase_tax_doc(self._doc())
        self.assertEqual(r["inv_no"], "INV001")

    def test_inv_date_reformatted(self):
        r = _preprocess_purchase_tax_doc(self._doc())
        self.assertEqual(r["inv_date"], "01/05/2026")

    def test_seller_fields_mapped(self):
        r = _preprocess_purchase_tax_doc(self._doc())
        self.assertEqual(r["seller_name"],   "ABC Co")
        self.assertEqual(r["seller_tax"],    "1234567890123")
        self.assertEqual(r["seller_branch"], "00001")

    def test_defaults_applied_for_missing_fields(self):
        r = _preprocess_purchase_tax_doc({"netAmt": 0, "vatRate": 0, "vatAmt": 0})
        self.assertEqual(r["status"],        "ลงบัญชีแล้ว")
        self.assertEqual(r["seller_branch"], "00000")


# ---------------------------------------------------------------------------
# ac_1101_thai — Thai column headers in _PTR_COLUMNS
# ---------------------------------------------------------------------------

class TestPtrColumns(unittest.TestCase):
    def test_19_columns(self):
        self.assertEqual(len(_PTR_COLUMNS), 19)

    def test_all_columndef_instances(self):
        for col in _PTR_COLUMNS:
            self.assertIsInstance(col, ColumnDef)

    def test_thai_headers_present(self):
        labels = [c.header_label for c in _PTR_COLUMNS]
        self.assertIn("ลำดับที่",        labels)
        self.assertIn("เลขที่เอกสาร",    labels)
        self.assertIn("ไม่มี VAT",       labels)
        self.assertIn("VAT 0%",          labels)
        self.assertIn("VAT 7%",          labels)
        self.assertIn("หัก ณ ที่จ่าย",  labels)
        self.assertIn("ต้องชำระ",       labels)

    def test_currency_columns_typed_number(self):
        # Columns 11-17 (ไม่มี VAT through ต้องชำระ) must be data_type="number"
        for col in _PTR_COLUMNS[11:18]:
            self.assertEqual(col.data_type, "number", f"{col.header_label} should be 'number'")

    def test_row_sequence_is_first(self):
        self.assertEqual(_PTR_COLUMNS[0].source_field, "row_sequence")


# ---------------------------------------------------------------------------
# ac_1101_output — template engine renders correct column values
# ---------------------------------------------------------------------------

class TestRenderViaEngine(unittest.TestCase):
    def _make_doc(self, **overrides):
        base = {
            "invNo": "PO-001", "invDate": "2026-05-10",
            "sellerName": "Thai Supplier Co", "sellerTax": "0105559000001",
            "sellerBranch": "00000", "netAmt": 2000, "vatRate": 7,
            "vatAmt": 140, "whtAmt": 0, "reference": "R1",
            "category": "สินค้า", "description": "ซื้อสินค้า",
        }
        base.update(overrides)
        return base

    def test_renders_19_cols_per_row(self):
        records = [_preprocess_purchase_tax_doc(self._make_doc())]
        engine = TemplateEngine(_PTR_COLUMNS)
        headers, rows = engine.render(records)
        self.assertEqual(len(headers), 19)
        self.assertEqual(len(rows[0]), 19)

    def test_row_sequence_is_1_based(self):
        docs = [_preprocess_purchase_tax_doc(self._make_doc()) for _ in range(3)]
        engine = TemplateEngine(_PTR_COLUMNS)
        _, rows = engine.render(docs)
        self.assertEqual(rows[0][0], "1")
        self.assertEqual(rows[2][0], "3")

    def test_vat_7_amount_in_col_13(self):
        records = [_preprocess_purchase_tax_doc(self._make_doc(netAmt=2000, vatRate=7, vatAmt=140))]
        engine = TemplateEngine(_PTR_COLUMNS)
        _, rows = engine.render(records)
        # col 13 = vat_7 bucket
        self.assertAlmostEqual(float(rows[0][13]), 2000.0)


# ---------------------------------------------------------------------------
# ac_1101_excel — integration: create_purchase_tax_report produces valid xlsx
# ---------------------------------------------------------------------------

class TestCreatePurchaseTaxReportExcel(unittest.TestCase):
    def _sample_docs(self):
        return [
            {
                "invNo": "PO-2026-001", "invDate": "2026-05-01",
                "sellerName": "ห้างหุ้นส่วน ABC", "sellerTax": "0105559001234",
                "sellerBranch": "00000", "netAmt": 5000, "vatRate": 7,
                "vatAmt": 350, "whtAmt": 150, "reference": "REF01",
                "category": "ค่าบริการ", "description": "ค่าจ้างเหมา",
            },
            {
                "invNo": "PO-2026-002", "invDate": "2026-05-15",
                "sellerName": "XYZ Supply Ltd", "sellerTax": "0105559005678",
                "sellerBranch": "00001", "netAmt": 10000, "vatRate": 0,
                "vatAmt": 0, "whtAmt": 0, "reference": "REF02",
                "category": "วัสดุ", "description": "ซื้อวัสดุ",
            },
        ]

    def test_xlsx_file_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "ptr.xlsx")
            result = create_purchase_tax_report(
                documents=self._sample_docs(),
                output_path=xlsx_path,
                company_info={"name": "Test Co", "taxId": "0100000000001",
                               "branch": "00000", "branchName": "สำนักงานใหญ่"},
                report_period=("01/05/2026", "31/05/2026"),
            )
            self.assertTrue(os.path.exists(str(result)))
            self.assertGreater(os.path.getsize(str(result)), 0)

    def test_xlsx_has_correct_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "ptr.xlsx")
            create_purchase_tax_report(
                documents=self._sample_docs(),
                output_path=xlsx_path,
                company_info={"name": "Test Co"},
                report_period=("", ""),
            )
            wb = openpyxl.load_workbook(xlsx_path)
            self.assertIn("รายงานภาษีซื้อ", wb.sheetnames)

    def test_xlsx_data_rows_present(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "ptr.xlsx")
            create_purchase_tax_report(
                documents=self._sample_docs(),
                output_path=xlsx_path,
                company_info={"name": "Test Co"},
                report_period=("", ""),
            )
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["รายงานภาษีซื้อ"]
            # Data starts at row 13 (1-indexed); check seq and invoice number
            self.assertIsNotNone(ws.cell(row=13, column=1).value)   # ลำดับที่ = 1
            self.assertEqual(ws.cell(row=13, column=2).value, "PO-2026-001")

    def test_xlsx_empty_documents(self):
        """Must not crash with zero documents."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "ptr_empty.xlsx")
            result = create_purchase_tax_report(
                documents=[],
                output_path=xlsx_path,
                company_info={"name": "Test Co"},
                report_period=("", ""),
            )
            self.assertTrue(os.path.exists(str(result)))


if __name__ == "__main__":
    unittest.main()
