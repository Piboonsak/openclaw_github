"""Unit tests for TemplateEngine (TASK-1001).

Covers all 9 acceptance criteria:
  ac_1001_map       — field mapping order
  ac_1001_transform — all 8 transforms
  ac_1001_date_text — CSV date-as-text (=""value"")
  ac_1001_express   — Express transaction fields (row_sequence, aliases)
  ac_1001_csv       — CSV encoding (utf-8, utf-8-bom, tis-620)
  ac_1001_excel     — Excel output (valid xlsx, styled headers)
  ac_1001_missing   — missing fields use default / empty (no crash)
  ac_1001_multi     — multiple documents rendered in row order
  ac_1001_product   — product fields graceful fallback until TASK-1013
"""

from __future__ import annotations

import csv
import io
import unittest

from src.backend.services.template_engine import (
    ColumnDef,
    TemplateEngine,
    _parse_ce_date,
    _to_thai_date_short,
    _to_thai_date_full,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _engine(*cols: ColumnDef, encoding: str = "utf-8") -> TemplateEngine:
    return TemplateEngine(list(cols), encoding=encoding)


def _col(source: str, label: str, dtype: str = "string", transform: str | None = None, default: str | None = None) -> ColumnDef:
    return ColumnDef(source_field=source, header_label=label, data_type=dtype, transform=transform, default_value=default)


def _read_csv(data: bytes, encoding: str = "utf-8") -> list[list[str]]:
    text = data.decode(encoding, errors="replace")
    return list(csv.reader(io.StringIO(text)))


# ---------------------------------------------------------------------------
# ac_1001_map — field mapping order
# ---------------------------------------------------------------------------

class TestFieldMappingOrder(unittest.TestCase):
    """ac_1001_map: output columns match template column order exactly."""

    def test_field_mapping_order(self):
        cols = [
            _col("voucher_no",   "Voucher"),
            _col("voucher_date", "Date"),
            _col("debit",        "Debit"),
            _col("credit",       "Credit"),
        ]
        engine = TemplateEngine(cols)
        record = {"voucher_no": "VN001", "voucher_date": "2026-05-01", "debit": "5000", "credit": "0"}
        headers, rows = engine.render([record])
        self.assertEqual(headers, ["Voucher", "Date", "Debit", "Credit"])
        self.assertEqual(rows[0][0], "VN001")
        self.assertEqual(rows[0][1], "2026-05-01")
        self.assertEqual(rows[0][2], "5000")
        self.assertEqual(rows[0][3], "0")

    def test_column_order_preserved_different_positions(self):
        cols = [_col("credit", "Credit"), _col("debit", "Debit"), _col("voucher_no", "Voucher")]
        engine = TemplateEngine(cols)
        headers, rows = engine.render([{"voucher_no": "X", "debit": "100", "credit": "0"}])
        self.assertEqual(headers[0], "Credit")
        self.assertEqual(headers[1], "Debit")
        self.assertEqual(headers[2], "Voucher")


# ---------------------------------------------------------------------------
# ac_1001_transform — all 8 transforms
# ---------------------------------------------------------------------------

class TestTransforms(unittest.TestCase):
    """ac_1001_transform: each of the 8 supported transforms produces correct output."""

    def _apply(self, value: str, transform: str, record: dict | None = None) -> str:
        engine = TemplateEngine([])
        return engine.apply_transform(value, transform, record or {})

    # 1. uppercase
    def test_uppercase(self):
        self.assertEqual(self._apply("hello world", "uppercase"), "HELLO WORLD")
        self.assertEqual(self._apply("บริษัท abc", "uppercase"), "บริษัท ABC")

    # 2. pad_left
    def test_pad_left_zeros(self):
        self.assertEqual(self._apply("510", "pad_left:5:0"), "00510")
        self.assertEqual(self._apply("51000", "pad_left:5:0"), "51000")

    def test_pad_left_already_full(self):
        self.assertEqual(self._apply("12345", "pad_left:5:0"), "12345")

    # 3. thai_date (backward-compat alias for thai_date_full)
    def test_thai_date_iso(self):
        result = self._apply("2026-05-01", "thai_date")
        self.assertEqual(result, "1/5/2569")

    # 4. strip_dash
    def test_strip_dash(self):
        self.assertEqual(self._apply("0105559-12-3456", "strip_dash"), "010555912-3456".replace("-", ""))
        self.assertEqual(self._apply("0105559123456", "strip_dash"), "0105559123456")

    def test_strip_dash_removes_all(self):
        self.assertEqual(self._apply("1-2-3-4", "strip_dash"), "1234")

    # 5. thai_date_short
    def test_thai_date_short_iso(self):
        # 2026 CE → BE 2569 → short "69"
        self.assertEqual(self._apply("2026-05-04", "thai_date_short"), "04/05/69")

    def test_thai_date_short_preserves_day_padding(self):
        self.assertEqual(self._apply("2026-01-07", "thai_date_short"), "07/01/69")

    # 6. thai_date_full
    def test_thai_date_full_iso(self):
        self.assertEqual(self._apply("2026-05-01", "thai_date_full"), "1/5/2569")

    def test_thai_date_full_day_single_digit(self):
        self.assertEqual(self._apply("2026-07-04", "thai_date_full"), "4/7/2569")

    # 7. prefix
    def test_prefix(self):
        self.assertEqual(self._apply("6905/001", "prefix:OE"), "OE6905/001")

    def test_prefix_empty_value(self):
        self.assertEqual(self._apply("", "prefix:OE"), "OE")

    # 8. doc_number
    def test_doc_number_slash_pattern(self):
        record = {"voucher_date": "2026-05-01", "row_sequence": "1"}
        # YYMM/NNN: YYMM=6905, NNN=001
        result = self._apply("", "doc_number:YYMM/NNN", record)
        self.assertEqual(result, "6905/001")

    def test_doc_number_hash_pattern(self):
        record = {"voucher_date": "2026-05-01", "row_sequence": "3"}
        # YYMM######: YYMM=6905, ######=000003 → 10 chars
        result = self._apply("", "doc_number:YYMM######", record)
        self.assertEqual(result, "6905000003")
        self.assertEqual(len(result), 10)

    def test_doc_number_uses_row_sequence(self):
        record = {"voucher_date": "2026-05-01", "row_sequence": "42"}
        result = self._apply("", "doc_number:YYMM/NNN", record)
        self.assertEqual(result, "6905/042")

    def test_no_transform_returns_unchanged(self):
        engine = TemplateEngine([])
        self.assertEqual(engine.apply_transform("hello", None), "hello")
        self.assertEqual(engine.apply_transform("hello", ""), "hello")


# ---------------------------------------------------------------------------
# ac_1001_date_text — CSV dates written as =""value"" for Excel compatibility
# ---------------------------------------------------------------------------

class TestCsvDateAsText(unittest.TestCase):
    """ac_1001_date_text: date columns wrapped with =""..."" in CSV output."""

    def test_date_column_wrapped_as_excel_text(self):
        cols = [
            _col("voucher_date", "วันที่", dtype="date", transform="thai_date_short"),
            _col("debit",        "เดบิต",  dtype="number"),
        ]
        engine = TemplateEngine(cols)
        records = [{"voucher_date": "2026-05-04", "debit": "5000"}]
        headers, rows = engine.render(records)
        csv_bytes = engine.write_csv(headers, rows, cols)
        lines = _read_csv(csv_bytes)
        # The date value should be wrapped as ="..."
        self.assertIn('="04/05/69"', lines[1][0])

    def test_non_date_column_not_wrapped(self):
        cols = [_col("debit", "เดบิต", dtype="number")]
        engine = TemplateEngine(cols)
        headers, rows = engine.render([{"debit": "5000"}])
        csv_bytes = engine.write_csv(headers, rows, cols)
        lines = _read_csv(csv_bytes)
        self.assertNotIn('="', lines[1][0])
        self.assertEqual(lines[1][0], "5000")

    def test_date_as_text_disabled(self):
        cols = [_col("voucher_date", "Date", dtype="date", transform="thai_date_short")]
        engine = TemplateEngine(cols)
        headers, rows = engine.render([{"voucher_date": "2026-05-04"}])
        csv_bytes = engine.write_csv(headers, rows, cols, date_as_excel_text=False)
        lines = _read_csv(csv_bytes)
        self.assertEqual(lines[1][0], "04/05/69")

    def test_date_column_wrapped_when_typed_as_string_but_has_date_transform(self):
        # HR-17-06 regression: the UI saves every column as data_type="string",
        # so relying on data_type alone left dates unwrapped and Excel mangled
        # them. A thai_date transform must still trigger the Excel-safe wrap.
        cols = [_col("invoice_date", "วันที่", dtype="string", transform="thai_date_short")]
        engine = TemplateEngine(cols)
        headers, rows = engine.render([{"invoice_date": "2026-05-04"}])
        csv_bytes = engine.write_csv(headers, rows, cols)
        lines = _read_csv(csv_bytes)
        self.assertIn('="04/05/69"', lines[1][0])

    def test_date_source_field_wrapped_even_without_transform(self):
        # A known date source field (invoice_date) with no transform and typed
        # as string must still be treated as a date for Excel safety.
        cols = [_col("invoice_date", "วันที่", dtype="string")]
        engine = TemplateEngine(cols)
        headers, rows = engine.render([{"invoice_date": "01/05/69"}])
        csv_bytes = engine.write_csv(headers, rows, cols)
        lines = _read_csv(csv_bytes)
        self.assertIn('="01/05/69"', lines[1][0])


# ---------------------------------------------------------------------------
# ac_1001_express — Express transaction fields
# ---------------------------------------------------------------------------

class TestExpressFields(unittest.TestCase):
    """ac_1001_express: row_sequence injected; Express field aliases resolve."""

    def test_row_sequence_auto_injected(self):
        cols = [_col("row_sequence", "ลำดับ")]
        engine = TemplateEngine(cols)
        records = [{"voucher_no": "X"}, {"voucher_no": "Y"}, {"voucher_no": "Z"}]
        _, rows = engine.render(records)
        self.assertEqual(rows[0][0], "1")
        self.assertEqual(rows[1][0], "2")
        self.assertEqual(rows[2][0], "3")

    def test_amount_before_tax_alias(self):
        cols = [_col("amount_before_tax", "จำนวนก่อนภาษี")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"net_amount": "10000"}])
        self.assertEqual(rows[0][0], "10000")

    def test_amount_including_tax_alias(self):
        cols = [_col("amount_including_tax", "รวมภาษี")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"total_amount": "10700"}])
        self.assertEqual(rows[0][0], "10700")

    def test_tax_invoice_number_alias(self):
        cols = [_col("tax_invoice_number", "เลขที่ใบกำกับ")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"invoice_number": "INV-001"}])
        self.assertEqual(rows[0][0], "INV-001")

    def test_formula_doc_number_from_record(self):
        cols = [_col("formula_doc_number", "เลขที่เอกสาร(สูตร)")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"formula_doc_number": "OE6905/001"}])
        self.assertEqual(rows[0][0], "OE6905/001")

    def test_document_number_via_transform(self):
        cols = [_col("row_sequence", "เลขที่", transform="doc_number:YYMM/NNN")]
        engine = TemplateEngine(cols)
        records = [{"voucher_date": "2026-05-01"}]
        _, rows = engine.render(records)
        self.assertEqual(rows[0][0], "6905/001")


# ---------------------------------------------------------------------------
# ac_1001_csv — CSV encoding
# ---------------------------------------------------------------------------

class TestCsvEncoding(unittest.TestCase):
    """ac_1001_csv: CSV output in utf-8, utf-8-bom, and tis-620."""

    THAI_HEADER = "วันที่"

    def _render_csv(self, encoding: str) -> bytes:
        cols = [_col("voucher_date", self.THAI_HEADER, dtype="date", transform="thai_date_short")]
        engine = TemplateEngine(cols, encoding=encoding)
        headers, rows = engine.render([{"voucher_date": "2026-05-04"}])
        return engine.write_csv(headers, rows, cols)

    def test_utf8_encoding(self):
        data = self._render_csv("utf-8")
        self.assertIn(self.THAI_HEADER.encode("utf-8"), data)

    def test_utf8_bom_encoding(self):
        data = self._render_csv("utf-8-bom")
        # UTF-8 BOM starts with EF BB BF
        self.assertEqual(data[:3], b"\xef\xbb\xbf")
        self.assertIn(self.THAI_HEADER.encode("utf-8"), data)

    def test_tis620_encoding(self):
        data = self._render_csv("tis-620")
        self.assertIn(self.THAI_HEADER.encode("cp874"), data)
        # Must NOT be valid UTF-8 (Thai TIS-620 bytes are not valid UTF-8)
        with self.assertRaises(UnicodeDecodeError):
            data.decode("utf-8")

    def test_delimiter_semicolon(self):
        cols = [_col("a", "A"), _col("b", "B")]
        engine = TemplateEngine(cols, delimiter=";")
        headers, rows = engine.render([{"a": "1", "b": "2"}])
        data = engine.write_csv(headers, rows, cols)
        line = data.decode("utf-8").splitlines()[1]
        self.assertIn(";", line)


# ---------------------------------------------------------------------------
# ac_1001_excel — Excel output
# ---------------------------------------------------------------------------

class TestExcelOutput(unittest.TestCase):
    """ac_1001_excel: Excel output is valid xlsx with styled headers."""

    def _engine_with_date_number(self) -> tuple[TemplateEngine, list[ColumnDef]]:
        cols = [
            _col("voucher_date", "Date",   dtype="date",   transform="thai_date_short"),
            _col("debit",        "Debit",  dtype="number"),
            _col("description",  "Desc",   dtype="string"),
        ]
        return TemplateEngine(cols), cols

    def test_excel_output_is_valid_xlsx(self):
        import openpyxl
        engine, cols = self._engine_with_date_number()
        records = [{"voucher_date": "2026-05-04", "debit": "5000", "description": "Test"}]
        headers, rows = engine.render(records)
        xlsx_bytes = engine.write_excel(headers, rows, cols)

        self.assertGreater(len(xlsx_bytes), 0)
        # Load back with openpyxl to verify structure
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Header row
        self.assertEqual(ws.cell(1, 1).value, "Date")
        self.assertEqual(ws.cell(1, 2).value, "Debit")
        self.assertEqual(ws.cell(1, 3).value, "Desc")
        # Data row
        self.assertEqual(ws.cell(2, 1).value, '="04/05/69"'.strip('="'))

    def test_excel_output_non_empty(self):
        engine, cols = self._engine_with_date_number()
        headers, rows = engine.render([{"voucher_date": "2026-05-01", "debit": "100", "description": "x"}])
        xlsx_bytes = engine.write_excel(headers, rows, cols)
        self.assertGreater(len(xlsx_bytes), 5000)  # should be a real xlsx, not empty


# ---------------------------------------------------------------------------
# ac_1001_missing — missing fields use default / empty, no crash
# ---------------------------------------------------------------------------

class TestMissingFields(unittest.TestCase):
    """ac_1001_missing: missing source fields fall back to default_value or ''."""

    def test_missing_field_uses_default_value(self):
        cols = [_col("vendor_code", "รหัสผู้จำหน่าย", default="UNKNOWN")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"voucher_no": "X"}])
        self.assertEqual(rows[0][0], "UNKNOWN")

    def test_missing_field_no_default_is_empty(self):
        cols = [_col("vendor_name", "ชื่อผู้จำหน่าย")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"voucher_no": "X"}])
        self.assertEqual(rows[0][0], "")

    def test_null_field_value_uses_default(self):
        cols = [_col("description", "คำอธิบาย", default="(none)")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"description": None}])
        self.assertEqual(rows[0][0], "(none)")

    def test_all_missing_no_crash(self):
        cols = [_col("x", "X"), _col("y", "Y"), _col("z", "Z")]
        engine = TemplateEngine(cols)
        # Should not raise even when record is empty
        _, rows = engine.render([{}])
        self.assertEqual(rows[0], ["", "", ""])


# ---------------------------------------------------------------------------
# ac_1001_multi — multiple documents rendered in correct row order
# ---------------------------------------------------------------------------

class TestMultiDocumentRender(unittest.TestCase):
    """ac_1001_multi: multiple records rendered in input order with correct seq."""

    def test_multi_document_row_order(self):
        cols = [
            _col("row_sequence", "Seq"),
            _col("voucher_no",   "Voucher"),
        ]
        engine = TemplateEngine(cols)
        records = [
            {"voucher_no": "A"},
            {"voucher_no": "B"},
            {"voucher_no": "C"},
        ]
        _, rows = engine.render(records)
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0], ["1", "A"])
        self.assertEqual(rows[1], ["2", "B"])
        self.assertEqual(rows[2], ["3", "C"])

    def test_row_sequence_resets_per_render_call(self):
        cols = [_col("row_sequence", "Seq")]
        engine = TemplateEngine(cols)
        _, rows1 = engine.render([{}, {}])
        _, rows2 = engine.render([{}, {}, {}])
        self.assertEqual(rows1[-1][0], "2")
        self.assertEqual(rows2[-1][0], "3")


# ---------------------------------------------------------------------------
# ac_1001_product — product fields graceful fallback until TASK-1013
# ---------------------------------------------------------------------------

class TestProductFieldGraceful(unittest.TestCase):
    """ac_1001_product: product fields return default or '' — no crash."""

    def test_product_code_graceful_empty(self):
        cols = [_col("product_code", "รหัสสินค้า")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{"voucher_no": "X"}])
        self.assertEqual(rows[0][0], "")

    def test_product_name_uses_default(self):
        cols = [_col("product_name", "ชื่อสินค้า", default="N/A")]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{}])
        self.assertEqual(rows[0][0], "N/A")

    def test_all_product_fields_no_crash(self):
        cols = [
            _col("product_code",       "รหัส"),
            _col("product_name",       "ชื่อ"),
            _col("product_unit",       "หน่วย"),
            _col("product_unit_price", "ราคา/หน่วย"),
        ]
        engine = TemplateEngine(cols)
        _, rows = engine.render([{}])
        self.assertEqual(rows[0], ["", "", "", ""])


# ---------------------------------------------------------------------------
# Helper: parse_ce_date + ColumnDef.from_dict
# ---------------------------------------------------------------------------

class TestHelpers(unittest.TestCase):
    def test_parse_ce_date_iso(self):
        d = _parse_ce_date("2026-05-01")
        self.assertEqual(d.year, 2026)
        self.assertEqual(d.month, 5)
        self.assertEqual(d.day, 1)

    def test_parse_ce_date_thai_short(self):
        # 04/05/69 = BE 2569 → CE 2026
        d = _parse_ce_date("04/05/69")
        self.assertEqual(d.year, 2026)

    def test_parse_ce_date_be_full(self):
        # 01/05/2569 → CE 2026
        d = _parse_ce_date("01/05/2569")
        self.assertEqual(d.year, 2026)

    def test_parse_ce_date_unparseable(self):
        self.assertIsNone(_parse_ce_date("not-a-date"))

    def test_column_def_from_dict(self):
        d = {
            "source_field": "voucher_date",
            "header_label": "Date",
            "data_type": "date",
            "transform": "thai_date_short",
        }
        col = ColumnDef.from_dict(d)
        self.assertEqual(col.source_field, "voucher_date")
        self.assertEqual(col.transform, "thai_date_short")
        self.assertIsNone(col.default_value)


if __name__ == "__main__":
    unittest.main()
